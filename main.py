import openpyxl

workbook = openpyxl.load_workbook("all_product_name.xlsx")
worksheet = workbook.worksheets[0]
uniqueSet = {''}

for row in worksheet.rows:
    for cell in row:
        a = cell.value
        if a == None:
            continue
        a = a.strip()
        uniqueSet.add(a)
    print()
print(uniqueSet)

worksheet2 = workbook.create_sheet()
worksheet2.title = "unique"
row = 1
col = 1

for val in uniqueSet:
    if val == "":
        continue
    worksheet2.cell(col, 1, val)
    col = col + 1

workbook.save(filename='product_name_unique_value.xlsx')
